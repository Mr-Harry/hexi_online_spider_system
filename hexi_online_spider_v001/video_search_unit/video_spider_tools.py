# -*- coding:utf-8 -*- #
# I don't like the world. I just like you
# author : pyl owo,
# time : 2020/7/21
import re
from _md5 import md5
from datetime import datetime

import openpyxl
import pymysql

from video_search_unit.Video_Infringement_Config import config_of_video as config


def get_proxy():
    proxy = "%(user)s:%(pass)s@%(host)s:%(port)s" % {
        "host": config["proxyHost"],
        "port": config["proxyPort"],
        "user": config["proxyUser"],
        "pass": config["proxyPass"],
    }
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy,
    }
    return proxies
def get_video_search_type_list(setting_key :str, video_search_type=''):
    if not video_search_type:
        video_search_type = '1'
    search_type_set = set()
    for i in video_search_type.split('_'):
        type_list = config.get(setting_key, {}).get(int(i))
        if type_list:
            search_type_set.update(set(type_list))
    return list(search_type_set)

def mysql_save_to_current_result_table(result):
    conn = pymysql.connect(host=config["mysql_host"], port=config["mysql_port"], user=config["mysql_username"],
                           passwd=config["mysql_userpwd"], db=config["mysql_db"],
                           charset='utf8mb4')
    cursor = conn.cursor()
    table_name = config["mysql_result_table_name"] + datetime.today().strftime("%Y%m%d")
    now_time = datetime.now().strftime("%Y-%m-%d")
    for each in result:
        # each["qingquan_flag"] = judge_song_type_easy(each) # 判断规则
        # print(each["qingquan_flag"])

        video_title = pymysql.escape_string(each.get("video_title")) if each.get("video_title") else ""
        video_author = pymysql.escape_string(each.get("video_author")) if each.get("video_author") else ""
        video2_title = pymysql.escape_string(each.get("video2_title")) if each.get("video2_title") else ""
        video2_author = pymysql.escape_string(each.get("video2_author")) if each.get("video2_author") else ""

        sql_save_info = "insert into {}(yangben_URL,yangben_title,yangben_author_name," \
                        "qinquan_title,qinquan_author_name,qinquan_URL,qinquan_url_hash,qinquan_platform," \
                        "yangben_platform,t,qinquan_type,qingquan_flag,t_timestamp) values('{}','{}','{}','{}'," \
                        "'{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(
            table_name,
            each.get("video_url", ""), video_title, video_author, video2_title, video2_author,
            each.get("video2_url", ""), each.get("video2_url_hash"),
            each.get("video2_platform", ""), each.get("video_platform", ""), now_time, 1, 0, now_time)
        print('执行语句：')
        print(sql_save_info)
        try:
            # print(sql_save_info)
            cursor.execute(sql_save_info)
            conn.commit()
        except Exception as e:
            # print(sql_save_info)
            print(e)
            print("重复插入错误")
            pass
        else:
            # print("无关的音乐->{} 样本音乐->{}".format(each["audio2_songName"],each["audio_title"]))
            pass


# Md5 加密函数 32 返回32位的加密结果
def md5_use(text: str) -> str:
    result = md5(bytes(text, encoding="utf-8")).hexdigest()
    # print(result)
    return result


def create_excel_public(title_dict: dict, list_data: list, sheet="sheet01"):
    if not (title_dict and list_data):
        return
    wb = openpyxl.Workbook()
    mySheet = wb.create_sheet(index=0, title=sheet)
    # for i in range(cols_num):
    #     mySheet.cell(row=1, column=i + 1).value = title_dict.get(arr[i])
    # 第一行标题生成
    for i, title_name in enumerate(title_dict.values()):
        mySheet.cell(row=1, column=i + 1).value = title_name
    # 表格内容生成
    for i, data_dict in enumerate(list_data):
        for j, data_key in enumerate(title_dict.keys()):
            try:
                mySheet.cell(row=i + 2, column=j + 1).value = data_dict.get(data_key)
            except:
                print(data_dict)
    file_path = "wangyiyun7wceshi.xlsx"
    try:
        wb.save(file_path)
    except:
        return
    else:
        return file_path


if __name__ == '__main__':
    get_proxy()
    exit(0)
    configg = {
    "mysql_port": 55306,
    "mysql_username": 'root',  # 用户名 本地是root Xueyiyang
    "mysql_userpwd": 'Xueyiyang',  # 用户密码
    # test
    "mysql_host": '121.196.126.218',  # 本地地址 localhost
    "mysql_db": "zi_jie_tiao_dong",  # 数据库名称
    # "mysql_task_qq_table":"audio_task2020",  # 音频任务表 qq音乐先用这个
    # "mysql_result_table_name":"result_data_normal_",  # 音频结果表 后面拼接当天的时间 20200702
    # "mysql_result_table_name_test":"", # 如果该配置为空 默认存到当天数据表格
    }
    # title_dict = {
    #     'audio_title': '歌曲名',
    #     'audio_url': '链接',
    #     'audio_author': '歌手名',
    #     'audio_album': '专辑',
    # }
    title_dict = {
        'yangben_platform': '样本平台名称',
        'yangben_author_name': '样本的作者名称',
        'yangben_title': '样本的标题',
        'yangben_text': '样本文章内容（图文：文章，音频：专辑，视频：视频信息）',
        'yangben_URL': '样本链接URL',
        'yangben_pub_time': '样本的发布时间',
        'qinquan_platform': '侵权平台名称',
        'qinquan_author_name': '侵权样本作者名称',
        'qinquan_title': '侵权样本标题',
        'qinquan_text': '侵权样本内容 （图文：文章，音频：专辑，视频：视频信息）',
        'qinquan_URL': '侵权样本链接',
        'qinquan_pub_time': '侵权样本发布时间',
        'yewu': '业务人员的对接',
        't': '任务发布的时间（入库时间 格式化）',
        'similar_number': '相似度',
        'qinquan_type': '侵权类型 （0 图文 1 视频 2 音频）',
        'qingquan_flag': '用于判断是否侵权的flag （1 侵权 其余不是）',
    }
    conn = pymysql.connect(host=configg["mysql_host"], port=configg["mysql_port"], user=configg["mysql_username"],
                           passwd=configg["mysql_userpwd"], db=configg["mysql_db"],
                           charset='utf8mb4')
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    print(11111)
    # sql = "select * from audio_task2020 where audio_platform_id=3 and audio_flag_int = 1"
#     sql = """
#     SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200724 where yangben_platform ="网易云七万" union
# SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200725 where yangben_platform ="网易云七万"
# """
    sql ="""
    SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200727 where yangben_platform = "网易云七万测试";
    """
    cursor.execute(sql)
    print(22222)
    data = cursor.fetchall()
    print(33333)
    print(len(data))
    cursor.close()
    conn.close()
    create_excel_public(title_dict, data)

    # # 情咖QQ音乐样本导出
    # title_dict = {
    #     'yangben_platform': '样本平台名称',
    #     'yangben_author_name': '样本的作者名称',
    #     'yangben_title': '样本的标题',
    #     'yangben_text': '样本文章内容（图文：文章，音频：专辑，视频：视频信息）',
    #     'yangben_URL': '样本链接URL',
    #     'yangben_pub_time': '样本的发布时间',
    #     'qinquan_platform': '侵权平台名称',
    #     'qinquan_author_name': '侵权样本作者名称',
    #     'qinquan_title': '侵权样本标题',
    #     'qinquan_text': '侵权样本内容 （图文：文章，音频：专辑，视频：视频信息）',
    #     'qinquan_URL': '侵权样本链接',
    #     'qinquan_pub_time': '侵权样本发布时间',
    #     'yewu': '业务人员的对接',
    #     't': '任务发布的时间（入库时间 格式化）',
    #     'similar_number': '相似度',
    #     'qinquan_type': '侵权类型 （0 图文 1 视频 2 音频）',
    #     'qingquan_flag': '用于判断是否侵权的flag （1 侵权 其余不是）',
    # }
    # conn = pymysql.connect(host=configg["mysql_host"], port=configg["mysql_port"], user=configg["mysql_username"],
    #                        passwd=configg["mysql_userpwd"], db=configg["mysql_db"],
    #                        charset='utf8mb4')
    # cursor = conn.cursor(pymysql.cursors.DictCursor)
    # # sql = "select * from result_data_normal_20200723"
    # sql = "SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200718 where qinquan_platform ='情咖FM' union " \
    #       "SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200719 where qinquan_platform ='情咖FM' union " \
    #       "SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200720 where qinquan_platform ='情咖FM' union " \
    #       "SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200721 where qinquan_platform ='情咖FM' union " \
    #       "SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200722 where qinquan_platform ='情咖FM' union " \
    #       "SELECT * FROM zi_jie_tiao_dong.result_data_normal_20200723 where qinquan_platform ='情咖FM'"
    # # name_qingka_list = ['YG ', '中唱', '孔雀廊', '寰亚', '杰威尔', '树音乐', '相信音乐', '福茂', '英皇', '鸟人']
    # cursor.execute(sql)
    # data = cursor.fetchall()
    # print(data)
    # cursor.close()
    # conn.close()
    # create_excel_public(title_dict, data)
